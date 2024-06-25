import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import openpyxl
import threading
import time
import json
from os import system

wb = Workbook()
ws = wb.active
ws.append(['Username', 'First Name', 'Last Name', 'Bio', "Gender"])

males = open("./names/male_names_rus.txt", "r", encoding='utf-8').read().split("\n")
females = open("./names/female_names_rus.txt", "r", encoding='utf-8').read().split("\n")
males += open("./names/male_names_eng.txt", "r", encoding='utf-8').read().split("\n")
females += open("./names/female_names_eng.txt", "r", encoding='utf-8').read().split("\n")

for i in range(len(males)):
    males[i] = males[i].lower()
for i in range(len(females)):
    females[i] = females[i].lower()

lock = threading.Lock()
cnt = 0
cnt2 = 0

def solve(user):
    try:
        global cnt
        global cnt2
        user = user.split("@")[-1]

        url = 'https://t.me/' + user
        response = requests.get(url)

        if response.status_code == 200:
            html_content = response.content

            # Создаем объект BeautifulSoup, передавая HTML контент и парсер
            soup = BeautifulSoup(html_content, 'html.parser')

            # Находим мета-тег с именем "og:title" и получаем его содержимое
            title = soup.find('meta', {'property': 'og:title'})['content']
            if "Telegram: Contact" in title:
                cnt += 1
                cnt2 += 1
                system("cls")
                print("""  
 _______ _____                                     _    
 |__   __/ ____|     /\                            | |   
    | | | |  __     /  \   ___ ___ ___  _   _ _ __ | |_  
    | | | | |_ |   / /\ \ / __/ __/ _ \| | | | '_ \| __| 
    | | | |__| |  / ____ \ (_| (_| (_) | |_| | | | | |_  
  __|_|  \_____| /_/    \_\___\___\___/ \__,_|_| |_|\__| 
 |  __ \      | |                                        
 | |  | | __ _| |_ __ _   _ __   __ _ _ __ ___  ___ _ __ 
 | |  | |/ _` | __/ _` | | '_ \ / _` | '__/ __|/ _ \ '__|
 | |__| | (_| | || (_| | | |_) | (_| | |  \__ \  __/ |   
 |_____/ \__,_|\__\__,_| | .__/ \__,_|_|  |___/\___|_|   v04
                         | |                             
                         |_|     
""")
                print("Аккаунтов просканировано:",cnt, "/", len(data))
                print("Неавлидных:", cnt2)
                return

            # То же самое для мета-тега с описанием
            description = soup.find('meta', {'property': 'og:description'})['content']
            if ("You can contact" in description):
                description = ""
            first_name = str(soup.find('div', {'class': 'tgme_page_additional'})).split("<strong>")[2].split("</strong>")[0]
            last_name = title.split(first_name)[1]
            data2 = "Неопределено"
            first_name_cp = ""
            for i in first_name:
                if ((i >= 'a' and i <= 'z') or (i >= 'A' and i <= 'Z') or i == ' ' or (i >= 'А' and i <= 'Я') or (i >= 'а' and i <= 'я')):
                    first_name_cp += i
            for i in first_name_cp.lower().split():
                if (i in males):
                    data2 = "М"
            for i in first_name_cp.lower().split():
                if (i in females):
                    data2 = "Ж"
            
            cnt += 1
            system("cls")
            print("""  
 _______ _____                                     _    
 |__   __/ ____|     /\                            | |   
    | | | |  __     /  \   ___ ___ ___  _   _ _ __ | |_  
    | | | | |_ |   / /\ \ / __/ __/ _ \| | | | '_ \| __| 
    | | | |__| |  / ____ \ (_| (_| (_) | |_| | | | | |_  
  __|_|  \_____| /_/    \_\___\___\___/ \__,_|_| |_|\__| 
 |  __ \      | |                                        
 | |  | | __ _| |_ __ _   _ __   __ _ _ __ ___  ___ _ __ 
 | |  | |/ _` | __/ _` | | '_ \ / _` | '__/ __|/ _ \ '__|
 | |__| | (_| | || (_| | | |_) | (_| | |  \__ \  __/ |   
 |_____/ \__,_|\__\__,_| | .__/ \__,_|_|  |___/\___|_|   v04
                         | |                             
                         |_|     
""")
            print("Аккаунтов просканировано:",cnt, "/", len(data))
            print("Неавлидных:", cnt2)
            # Acquire lock before writing to the worksheet
            with lock:
                ws.append(["@" + user, first_name_cp, last_name, description, data2])
        else:
            print("Failed to retrieve the webpage")
    except:
        pass
def solve_async(i):
    thread = threading.Thread(target=solve, args=(i,))
    thread.start()
    return thread

def filter_xlsx():
    try:
        file = input("Введите путь до xlsx файла: ")
        keys = input("Введите ключевые слова через запятую (например футбол,кот,собака): ").split(",")
        wb_obj = openpyxl.load_workbook(file)
        sheet_obj = wb_obj.active
        
        for i in range (2, sheet_obj.max_row + 1):
            for e in range(1, 5):
                cell_obj = sheet_obj.cell(row = i, column = e)
                if (cell_obj.value):
                    for u in keys:
                        if (u.lower() in cell_obj.value.lower()):
                            ws.append([sheet_obj.cell(row = i, column = 1).value, sheet_obj.cell(row = i, column = 2).value, sheet_obj.cell(row = i, column = 3).value, sheet_obj.cell(row = i, column = 4).value, sheet_obj.cell(row = i, column = 5).value])
                            break
        wb.save("user_info_filtered.xlsx")
        print("Задание успешно выполнено! Данные сохранены в user_info_filtered.xlsx")
        input("Для выхода нажмите Enter...")
    except:
        pass

print("""  
 _______ _____                                     _    
 |__   __/ ____|     /\                            | |   
    | | | |  __     /  \   ___ ___ ___  _   _ _ __ | |_  
    | | | | |_ |   / /\ \ / __/ __/ _ \| | | | '_ \| __| 
    | | | |__| |  / ____ \ (_| (_| (_) | |_| | | | | |_  
  __|_|  \_____| /_/    \_\___\___\___/ \__,_|_| |_|\__| 
 |  __ \      | |                                        
 | |  | | __ _| |_ __ _   _ __   __ _ _ __ ___  ___ _ __ 
 | |  | |/ _` | __/ _` | | '_ \ / _` | '__/ __|/ _ \ '__|
 | |__| | (_| | || (_| | | |_) | (_| | |  \__ \  __/ |   
 |_____/ \__,_|\__\__,_| | .__/ \__,_|_|  |___/\___|_|    v04
                         | |                             
                         |_|     
""")

move = input("(1) Провести поиск по ключевым словам\n(2) Спарсить данные тг аккаунтов\nВыберите действие: ")
if (move == "1"):
    filter_xlsx()
    exit()
path = input("Введите путь до базы юзернеймов (например accounts.txt): ")
sleept = float(input("Введите задержку между запросами в секундах (например 0.25, меньше база - меньше можно ставить задержку): "))

data = open(path, "r").read().split("\n")

# List to store threads
threads = []

for i in data:
    time.sleep(sleept)
    thread = solve_async(i)
    threads.append(thread)

# Wait for all threads to complete
for thread in threads:
    
    thread.join()

wb.save("user_info.xlsx")
print("Задание успешно выполнено! Данные сохранены в user_info.xlsx")
input("Для выхода нажмите Enter...")